#!/usr/bin/env python
# coding: utf-8

# In[11]:


import csv
import openpyxl
import os.path
import xlsxwriter 
import os
def main():
    line=list()
    workbook = xlsxwriter.Workbook('C:\\Users\\Mandy\\Desktop\\country\\Tripdotcom\\All.xlsx') 
    worksheet = workbook.add_worksheet() 
    workbook.close() 
    write_wb=openpyxl.load_workbook('C:\\Users\\Mandy\\Desktop\\country\\Tripdotcom\\All.xlsx')#엑셀 파일 열기
    write_ws=write_wb.active
    filenames=os.listdir('C:\\Users\\Mandy\\Desktop\\Tripdotcom')
    for filename in filenames:    
        f=open('C:\\Users\\Mandy\\Desktop\\Tripdotcom\\'+filename, 'r',encoding='utf-8')
        rdr = csv.reader(f)
        for a in rdr:
            line.append(a)     
        f.close()   
    write_ws.append(['TripDotCom','나라','도시','상품번호','상품명',
                   '옵션','URL','통화','가격','후기수','비고'])
    for item in line:
        write_ws.append(item) #ALL목록에 추가
        flag=0  #시트 목록에 이미 있는 나라인지 확인하는 플래그
        flag2=0
        for sheet in write_wb:
            if str(sheet)==str('<Worksheet "'+item[1]+'">'):#기존에 있는거면
                flag=1
                sheet.append(item)
            elif str(sheet)==str('<Worksheet "'+item[2]+'">'):
                flag2=1
                sheet.append(item)
        if flag==0: #없으면 시트 추가
            sheet2=write_wb.create_sheet(item[1])
            sheet2.append(['TripDotCom','나라','도시','상품번호','상품명',
                   '옵션','URL','통화','가격','후기수','비고'])
            sheet2.append(item)
        if flag2==0: #없으면 시트 추가
            sheet3=write_wb.create_sheet(item[2])
            sheet3.append(['TripDotCom','나라','도시','상품번호','상품명',
                   '옵션','URL','통화','가격','후기수','비고'])
            sheet3.append(item)    

    write_wb.save('C:\\Users\\Mandy\\Desktop\\country\\Tripdotcom\\All.xlsx')
    print('finish')
main()    


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




