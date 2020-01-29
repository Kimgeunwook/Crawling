#!/usr/bin/env python
# coding: utf-8

# In[9]:


import csv
import openpyxl
import os.path
import xlsxwriter 
import os

line=list()
workbook = xlsxwriter.Workbook('C:\\Users\\Mandy\\Desktop\\country\\Klook\\Klook상품크롤링.xlsx') 
worksheet = workbook.add_worksheet() 
workbook.close() 
write_wb=openpyxl.load_workbook('C:\\Users\\Mandy\\Desktop\\country\\Klook\\Klook상품크롤링.xlsx')#엑셀 파일 열기
write_ws=write_wb.active
for i in range(1,35000):
    if os.path.exists('C:\\Users\\Mandy\\Desktop\\Klook\\' + str(i) + '.csv'):
        f = open('C:\\Users\\Mandy\\Desktop\\Klook\\'+str(i)+'.csv', 'r',encoding='utf-8')
        rdr = csv.reader(f)
        for a in rdr:
            line.append(a)
        f.close()
write_ws.append(['Klook','나라','도시','상품번호','상품명',
               '옵션명','상품 URL','통화','가격','후기수','비고1','비고2'])
for item in line:
    write_ws.append(item) #ALL목록에 추가
    print(item)
    flag=0  #시트 목록에 이미 있는 나라인지 확인하는 플래그
    for sheet in write_wb:
        if str(sheet)==str('<Worksheet "'+item[1]+'">'):#기존에 있는거면
            flag=1
            sheet.append(item)
    if flag==0: #없으면 시트 추가
        sheet2=write_wb.create_sheet(item[1])
        sheet2.append(['Klook','나라','도시','상품번호','상품명',
               '옵션명','상품 URL','통화','가격','후기수','비고1','비고2'])
        sheet2.append(item)
        
write_wb.save('C:\\Users\\Mandy\\Desktop\\country\\Klook\\Klook상품크롤링.xlsx')
print('finish')    
    


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




